
# import pandas as pd
# from openpyxl import load_workbook

# # Replace 'your_template.xlsx' with the actual path to your Excel template
# template_file_path = '/Users/matthewmurphy/Documents/development/daily_planner_project/schedule_template.xlsx'
# excel_file_path = '/Users/matthewmurphy/Documents/development/daily_planner_project/Week 46 Daily Planner.xltx'

# # Step 1: Read CSV Data
# df = pd.read_csv('/Users/matthewmurphy/Documents/development/daily_planner_project/schedule_testv3.csv')

# df['Day'] = pd.to_datetime(df['Date']).dt.day_name()
# # Convert 'Shift Start Date Time' to datetime and format it
# df['Shift Start Date Time'] = pd.to_datetime(df['Shift Start Date Time'], format='%m/%d/%Y, %I:%M:%S %p')
# df['Shift Start Date Time'] = df['Shift Start Date Time'].dt.strftime('%I:%M %p')

# df['Shift End Date Time'] = pd.to_datetime(df['Shift End Date Time'], format='%m/%d/%Y, %I:%M:%S %p')
# df['Shift End Date Time'] = df['Shift End Date Time'].dt.strftime('%I:%M %p')

# # Step 2: Sort DataFrame
# df_sorted = df.sort_values(by=['Day', 'Employee Full Name', 'Shift Start Date Time'])
# print(df_sorted)
# # Define the custom order for the job groups
# custom_order = ['SM', 'SD', 'ICM', 'CSS', 'CSA', 'SOS', 'SOA', 'VMS']

# # Specify the columns to write to the Excel file
# columns_to_write = ['Employee Full Name', 'Shift Start Date Time', 'Shift End Date Time']

# # Drop rows where 'Shift Start Date Time' is missing or not a valid datetime
# df_sorted = df_sorted.dropna(subset=['Shift Start Date Time'])

# # Convert 'Shift Start Date Time' to datetime and format to hh:mm
# df_sorted['Shift Start Date Time'] = pd.to_datetime(df_sorted['Shift Start Date Time'], errors='coerce', format='%I:%M %p').dt.strftime('%I:%M %p')
# df_sorted['Shift End Date Time'] = pd.to_datetime(df_sorted['Shift End Date Time'], errors='coerce', format='%I:%M %p').dt.strftime('%I:%M %p')

# # Sort the DataFrame by custom order and then by 'Shift Start Date Time'
# df_sorted = df_sorted.sort_values(
#     by=['EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time'],
#     key=lambda x: pd.Categorical(x['EMP_COMMON_PRIMARY_JOB_1'], categories=custom_order, ordered=True)
# ).sort_values(by='Shift Start Date Time', key=lambda x: pd.to_datetime(x, format='%I:%M %p', errors='coerce').dt.time)

# # Drop duplicates based on relevant columns
# df_no_duplicates = df_sorted.drop_duplicates(subset=['EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time'])

# # Load the existing workbook
# book = load_workbook('/Users/matthewmurphy/Documents/development/daily_planner_project/Week 46 Daily Planner test1 copy 3.xlsx')

# # Get the sheet you want to append the data to (modify 'Sheet1' to the actual sheet name)
# writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')
# writer.book = book
# writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

# # Append the new data to the existing sheet starting at cell A3
# df_no_duplicates[columns_to_write].to_excel(writer, index=False, header=False, sheet_name='Sheet1', startrow=2, startcol=0)

# # Save the changes
# writer.save()


# Save the changes to the workbook
#writer.save()
################################################
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import NamedStyle

# # Replace 'your_file.csv' and 'your_workbook.xlsx' with the actual paths to your files
# csv_file_path = '/Users/matthewmurphy/Documents/development/daily_planner_project/schedule_testv3.csv'
# workbook_path = '/Users/matthewmurphy/Documents/development/daily_planner_project/Week 46 Daily Planner test1 copy 3.xlsx'

# # Step 1: Read CSV Data
# df = pd.read_csv(csv_file_path)

# # Assuming 'Shift Start Date Time' and 'Shift End Date Time' are datetime objects
# df['Day'] = pd.to_datetime(df['Date']).dt.day_name()
# df['Shift Start Date Time'] = pd.to_datetime(df['Shift Start Date Time'], format='%m/%d/%Y, %I:%M:%S %p')
# df['Shift End Date Time'] = pd.to_datetime(df['Shift End Date Time'], format='%m/%d/%Y, %I:%M:%S %p')

# # Format 'Shift Start Date Time' and 'Shift End Date Time' as 'HH:mm-HH:mm'
# df['Shift Time Range'] = df['Shift Start Date Time'].dt.strftime('%I:%M') + '-' + df['Shift End Date Time'].dt.strftime('%I:%M')

# # Create a Pandas Excel writer using Openpyxl as the engine
# with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a') as writer:
#     # Load the existing workbook
#     existing_workbook = writer.book
#     writer.sheets = dict((ws.title, ws) for ws in existing_workbook.worksheets)


#     # Iterate over each day and append data to the existing sheet
#     for day in df['Day'].unique():
#         # Select only rows for the current day
#         df_day = df[df['Day'] == day]

#         # Group by employee and aggregate shifts
#         grouped = df_day.groupby('Employee Full Name').agg({'Shift Time Range': '-'.join}).reset_index()

#         # Shorten employee names to first names
#         grouped['Employee First Name'] = grouped['Employee Full Name'].str.split().str[0]

#         # Write the data to the Excel sheet, starting from A3
#         sheet_name = str(day)
#         grouped.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2, startcol=0, header=False)

#         # Set column widths
#         current_sheet = writer[sheet_name]
#         current_sheet.column_dimensions['A'].width = 15
#         current_sheet.column_dimensions['B'].width = 20

# # Save the Excel file
# writer.save()
##########################################

# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Alignment

# # Replace 'your_file.csv' and 'your_workbook.xlsx' with the actual paths to your files
# csv_file_path = '/Users/matthewmurphy/Documents/development/daily_planner_project/schedule_testv3.csv'
# workbook_path = '/Users/matthewmurphy/Documents/development/daily_planner_project/Week 46 Daily Planner test1 copy 4.xlsx'

# # Step 1: Read CSV Data
# df = pd.read_csv(csv_file_path)

# df['Day'] = pd.to_datetime(df['Date']).dt.day_name()
# # Convert 'Shift Start Date Time' to datetime and format it
# df['Shift Start Date Time'] = pd.to_datetime(df['Shift Start Date Time'], format='%m/%d/%Y, %I:%M:%S %p')
# df['Shift Start Date Time'] = df['Shift Start Date Time'].dt.strftime('%I:%M %p')

# df['Shift End Date Time'] = pd.to_datetime(df['Shift End Date Time'], format='%m/%d/%Y, %I:%M:%S %p')
# df['Shift End Date Time'] = df['Shift End Date Time'].dt.strftime('%I:%M %p')

# # Step 2: Sort DataFrame
# df_sorted = df.sort_values(by=['Day', 'EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time', 'Employee Full Name'])

# # Open the existing workbook
# workbook = load_workbook(workbook_path)

# # Iterate over each day and append data to the corresponding sheet
# for day in df_sorted['Day'].unique():
#     # Select only rows for the current day
#     df_day = df_sorted[df_sorted['Day'] == day]

#     # Group by job title and shift start time, and sort each group
#     df_day_sorted = df_day.sort_values(by=['EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time'])

#     # Drop duplicates within the day based on relevant columns
#     df_day_no_duplicates = df_day_sorted.drop_duplicates(subset=['EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time'])

#     df_day_no_duplicates = df_day_no_duplicates.assign(
#         **{
#             'Employee First Name': df_day_no_duplicates['Employee Full Name'].apply(lambda x: x.split()[0]),
#             'Shift Times': df_day_no_duplicates['Shift Start Date Time'].astype(str) + '-' + df_day_no_duplicates['Shift End Date Time'].astype(str)
#         }
#     )

#     # Get the corresponding sheet in the workbook
#     sheet_name = str(day)
#     if sheet_name in workbook.sheetnames:
#         sheet_index = workbook.sheetnames.index(sheet_name)
#         sheet = workbook[workbook.sheetnames[sheet_index]]
#         print(sheet)
#     else:
#         sheet = workbook.create_sheet(title=sheet_name)
#         print(sheet.title + ' - this is a new sheet')


#     # Find the starting row in the sheet
#     start_row = 3
#     # start_row = sheet.max_row + 1 if sheet.max_row else 3

#     # Iterate over rows and append data to the sheet
#     # Iterate over rows and append data to the sheet
#     for r_idx, row in enumerate(df_day_no_duplicates.iterrows(), start_row):
#         col_employee_first_name = 'A'  # Column A starting from A3
#         if col_employee_first_name is not None:
#             sheet[col_employee_first_name + str(r_idx + 2)].value = row[1]['Employee First Name']

#         col_shift_times = 'B'  # Column B starting from B3
#         if col_shift_times is not None:
#             sheet[col_shift_times + str(r_idx + 2)].value = row[1]['Shift Times']
#             sheet[col_shift_times + str(r_idx + 2)].alignment = Alignment(horizontal='left')


# # Save the updated workbook
# workbook.save(workbook_path)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, range_boundaries

# Replace 'your_file.csv' and 'your_workbook.xlsx' with the actual paths to your files
csv_file_path = '/Users/matthewmurphy/Documents/development/daily_planner_project/schedule_testv3.csv'
workbook_path = '/Users/matthewmurphy/Documents/development/daily_planner_project/Week 46 Daily Planner1 copy 3.xlsx'
# Step 1: Read CSV Data
df = pd.read_csv(csv_file_path)

# Handle missing values


df['Day'] = pd.to_datetime(df['Date']).dt.day_name()
df = df.dropna(subset=['Day', 'EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time'])
# Convert 'Shift Start Date Time' to datetime and format it
df['Shift Start Date Time'] = pd.to_datetime(df['Shift Start Date Time'], format='%m/%d/%Y, %I:%M:%S %p')
df['Shift Start Date Time'] = df['Shift Start Date Time'].dt.strftime('%I:%M %p')

df['Shift End Date Time'] = pd.to_datetime(df['Shift End Date Time'], format='%m/%d/%Y, %I:%M:%S %p')
df['Shift End Date Time'] = df['Shift End Date Time'].dt.strftime('%I:%M %p')

# Sort DataFrame
df_sorted = df.sort_values(by=['Day', 'EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time', 'Employee Full Name'])

# Open the existing workbook
workbook = load_workbook(workbook_path)
print('Hello you big bitch')
print(workbook.sheetnames)
# Iterate over each day and append data to the corresponding sheet
for day in df_sorted['Day'].unique():
    # Select only rows for the current day
    df_day = df_sorted[df_sorted['Day'] == day]

    # Group by job title and shift start time, and sort each group
    df_day_sorted = df_day.sort_values(by=['EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time'])

    
    # Drop duplicates within the day based on relevant columns
    df_day_no_duplicates = df_day_sorted.drop_duplicates(subset=['Day', 'EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time'])
    
    # Define the custom order for the job groups
    job_groups_order = ['SD', 'SM', 'CSS', 'ICM', 'CSA', 'SOS', 'SOA', 'VMS']

    # Group by job title and shift start time, and sort each group
    print(df_day.columns)
    df_day_sorted = df_day.sort_values(by=['Day', pd.Categorical(df_day['EMP_COMMON_PRIMARY_JOB_1'], categories=job_groups_order, ordered=True), 'Shift Start Date Time'])




    # Drop duplicates within the day based on relevant columns
    df_day_no_duplicates = df_day_sorted.drop_duplicates(subset=['Day', 'EMP_COMMON_PRIMARY_JOB_1', 'Shift Start Date Time'])


    
    print(df_day_no_duplicates)
    df_day_no_duplicates = df_day_no_duplicates.assign(
        **{
            'Employee First Name': df_day_no_duplicates['Employee Full Name'].apply(lambda x: x.split()[0]),
            'Shift Times': df_day_no_duplicates['Shift Start Date Time'].astype(str) + '-' + df_day_no_duplicates['Shift End Date Time'].astype(str)
        }
    )
    
    # Get the corresponding sheet in the workbook
    sheet_name = str(day)
    if sheet_name in workbook.sheetnames:
        sheet_index = workbook.sheetnames.index(sheet_name)
        sheet = workbook[workbook.sheetnames[sheet_index]]
        print(sheet)
    else:
        sheet = workbook.create_sheet(title=sheet_name)
        print(sheet.title + ' - this is a new sheet')

    # Find the starting row in the sheet
    start_row = 3

    # Iterate over rows and append data to the sheet
    # Iterate over rows and append data to the sheet
    # Iterate over rows and append data to the sheet
    # Iterate over rows and append data to the sheet
    # Iterate over rows and append data to the sheet
    # Iterate over rows and append data to the sheet, starting from row 3
    # Iterate over rows and append data to the sheet
    for r_idx, row in enumerate(df_day_no_duplicates.iterrows(), start_row):
        col_employee_first_name = 'A'  # Column A starting from A3
        if col_employee_first_name is not None:
            cell_employee_first_name = sheet[col_employee_first_name + str(r_idx)]
            cell_employee_first_name.value = row[1]['Employee First Name']

        col_shift_times = 'B'  # Column B starting from B3
        if col_shift_times is not None:
            cell_shift_times = sheet[col_shift_times + str(r_idx)]

            # Check if the cell is part of a merged range
            is_merged = any(cell_shift_times.coordinate in merged_range for merged_range in sheet.merged_cells.ranges)

            # Handle merged cell differently, for example, by setting the value to the first cell in the range
            if is_merged:
                min_row, min_col, max_row, max_col = range_boundaries(cell_shift_times.coordinate)
                sheet[get_column_letter(min_col) + str(min_row)].value = row[1]['Shift Times']
                sheet[get_column_letter(min_col) + str(min_row)].alignment = Alignment(horizontal='left')
            else:
                # Set the value for non-merged cell
                cell_shift_times.value = row[1]['Shift Times']
                cell_shift_times.alignment = Alignment(horizontal='left')








# Save the updated workbook
workbook.save(workbook_path)
